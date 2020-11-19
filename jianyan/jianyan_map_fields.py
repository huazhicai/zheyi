from openpyxl import load_workbook
import difflib
import pymongo
import Levenshtein
from lxml import etree

from jianyan.excel_to_json import is_odd

host = 'localhost'
client = pymongo.MongoClient(host)
db = client.hemodialysis
collection = db.origin


def read_excel_data(sheet):
    for i in sheet['Q']:
        if i.value and i.value.startswith('zhyl151') and i.value != 'zhyl1510000':
            yield sheet['A%d' % i.row].value, sheet['B%d' % i.row].value, sheet['G%d' % i.row].value


def get_equal_rate(str1, str2):
    return difflib.SequenceMatcher(lambda x: x in " \\t", str1, str2).quick_ratio()


def leven_equal_rate(str1, str2):
    return Levenshtein.ratio(str1, str2)


def get_no_match_fields():
    file_handle = open("Jianyan.html", "r")  # 读取文件
    f = file_handle.read()
    file_handle.close()
    doc = etree.HTML(f)
    xiangmu = doc.xpath('/html/body/table[2]/tr/td[1]/text()')
    danwei = doc.xpath('/html/body/table[2]/tr/td[2]/text()')

    file_name = '../config/xuetou.xlsx'
    sheet_name = 'Sheet2'
    work_book = load_workbook(file_name)
    work_sheet = work_book[sheet_name]
    match_fields = []
    for row in work_sheet.values:
        if row[3]:
            match_fields.extend(list(filter(lambda x: x and is_odd(row.index(x)), row)))

    return [(name, unit) for name, unit in zip(xiangmu, danwei) if name not in set(match_fields)]


def save_no_match_fields():
    fields = get_no_match_fields()
    file_name = 'unmatch.xlsx'
    sheet_name = 'Sheet1'
    work_book = load_workbook(file_name)
    work_sheet = work_book[sheet_name]
    for data in fields:
        work_sheet.append(data)
    work_book.save(file_name)


def main(sheet, sheet2):
    file_handle = open("Jianyan.html", "r")  # 读取文件
    f = file_handle.read()
    file_handle.close()
    doc = etree.HTML(f)
    xiangmu = doc.xpath('/html/body/table[2]/tr/td[1]/text()')
    danwei = doc.xpath('/html/body/table[2]/tr/td[2]/text()')
    assert len(xiangmu) == len(danwei)

    for key, obj_field, obj_unit in read_excel_data(sheet):
        data = []
        unit_s = []
        for field, unit in zip(xiangmu, danwei):
            rate = get_equal_rate(field, obj_field.replace(" ", ""))
            if 0.8 < rate:
                # if 0.8 < rate < 0.88:
                if obj_unit and unit == obj_unit:
                    print(rate)
                    print(obj_field, obj_unit, ':', field, unit)
                    data.append(field)
                    unit_s.append(unit)
                if not obj_unit:
                    print(rate)
                    print(obj_field, obj_unit, ':', field, unit)
                    data.append(field)
                    unit_s.append(unit)
        data = sorted(data, key=lambda x: difflib.SequenceMatcher(None, x, obj_field).ratio(), reverse=True)
        new_data = []
        for i, val in enumerate(data):
            if unit_s[i] != '(null)':
                new_data.append(val)
                new_data.append(unit_s[i])
            else:
                new_data.append(val)
                new_data.append("")

        new_data.insert(0, obj_unit)
        new_data.insert(0, obj_field)
        new_data.insert(0, key)
        sheet2.append(new_data)


if __name__ == '__main__':
    file_name = 'xuetou.xlsx'
    sheet_name = '实验室检验信息'
    sheet_save = 'matchFields'
    work_book = load_workbook(file_name)
    work_sheet = work_book[sheet_name]
    sheet2 = work_book[sheet_save]
    main(work_sheet, sheet2)
    work_book.save(file_name)
